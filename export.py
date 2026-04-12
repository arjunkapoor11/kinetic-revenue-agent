import psycopg2
import requests
from dotenv import load_dotenv
import os
import statistics
import calendar
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.comments import Comment

load_dotenv()

TICKERS = [
    "SNOW", "DDOG", "MDB", "TENB", "QLYS",
    "NOW", "ADBE", "INTU", "WDAY", "PANW",
    "CDNS", "SNPS", "ADSK", "APP", "FTNT",
    "TEAM", "VEEV", "ROP", "PLTR",
    "HUBS", "ZS", "CRWD", "OKTA", "TTD",
    "GTLB", "BILL", "MNDY", "CFLT", "ESTC",
    "FROG", "S", "ZI", "DT", "TOST",
    "PCTY", "PAYC", "GWRE", "BSY", "CWAN",
    "NCNO", "BRZE", "KVYO", "PCOR", "SPSC",
    "MANH", "AZPN", "APPN", "DOCN",
]

DATA_COL = 4  # column D — first data column


# ── analytics (mirrors agent.py) ─────────────────────────────────────────

def get_db_connection():
    return psycopg2.connect(
        host=os.getenv("DB_HOST"), database=os.getenv("DB_NAME"),
        user=os.getenv("DB_USER"), password=os.getenv("DB_PASSWORD"),
        port=5432, sslmode="require")


def quarter_from_date(s):
    return (datetime.strptime(s, "%Y-%m-%d").month - 1) // 3 + 1


def quarter_end_date(y, q):
    m = q * 3
    return f"{y}-{m:02d}-{calendar.monthrange(y, m)[1]:02d}"


def next_period(period_str, steps=1):
    """Advance a period date by `steps` quarters (3 months each).

    Preserves fiscal calendar: uses the last day of the target month,
    so SNOW's Jan-31 advances to Apr-30, Jul-31, Oct-31, Jan-31.
    """
    d = datetime.strptime(period_str, "%Y-%m-%d")
    for _ in range(steps):
        m = d.month + 3
        y = d.year
        if m > 12:
            m -= 12
            y += 1
        d = datetime(y, m, calendar.monthrange(y, m)[1])
    return d.strftime("%Y-%m-%d")


def get_db_data(ticker):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT period,revenue FROM revenue_actuals WHERE ticker=%s ORDER BY period", (ticker,))
    actuals = [{"period": str(r[0]), "revenue": r[1]} for r in cur.fetchall()]
    cur.execute("SELECT period,estimated_revenue FROM consensus_estimates WHERE ticker=%s ORDER BY period", (ticker,))
    estimates = [{"period": str(r[0]), "estimated_revenue": r[1]} for r in cur.fetchall()]
    cur.close()
    conn.close()
    return actuals, estimates


def get_transcript_analyses(ticker):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT period, transcript_analysis FROM transcripts "
        "WHERE ticker=%s AND transcript_analysis IS NOT NULL", (ticker,))
    results = {str(r[0]): r[1] for r in cur.fetchall()}
    cur.close()
    conn.close()
    return results


def get_pre_earnings_consensus(ticker):
    """Read pre-earnings consensus (from FMP /earnings) keyed by period."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT period, estimated_revenue FROM pre_earnings_consensus "
        "WHERE ticker=%s ORDER BY period", (ticker,))
    results = {str(r[0]): r[1] for r in cur.fetchall()}
    cur.close()
    conn.close()
    return results


def compute_qoq(actuals):
    out = []
    for i in range(1, len(actuals)):
        prev, cur = actuals[i - 1]["revenue"], actuals[i]["revenue"]
        q = quarter_from_date(actuals[i]["period"])
        out.append({"period": actuals[i]["period"], "quarter": f"Q{q}",
                     "revenue": cur, "qoq_dollar_change": cur - prev,
                     "qoq_pct_change": round((cur - prev) / prev * 100, 2) if prev else 0})
    return out


def compute_seasonality(qoq):
    """Company-specific seasonal baselines from last 8 quarters per season."""
    by_q = defaultdict(list)
    for r in qoq:
        by_q[r["quarter"]].append(r["qoq_dollar_change"])
    s = {}
    for q in ("Q1", "Q2", "Q3", "Q4"):
        all_c = by_q.get(q, [])
        if not all_c:
            continue
        c = all_c[-8:]
        n = len(c)
        avg = statistics.mean(c)
        std = statistics.stdev(c) if n > 1 else 0
        cv = abs(std / avg) if avg != 0 else 0
        if cv > 0.4 and n >= 2:
            decay = 0.85
            weights = [decay ** (n - 1 - i) for i in range(n)]
            w_avg = sum(v * w for v, w in zip(c, weights)) / sum(weights)
        else:
            w_avg = avg
        s[q] = {"avg_qoq_change": round(w_avg),
                 "std_qoq_change": round(std),
                 "observations": n,
                 "cv": round(cv, 3),
                 "weighting": "exponential_decay" if cv > 0.4 and n >= 2 else "equal"}
    return s


def classify_seasonal_trend(vals, cv=0.0, pct_values=None):
    """Returns (trend, projected_value, is_pct_rate).
    Growing/accelerating/decelerating return a % rate; others return $ QoQ."""
    if not vals:
        return "no_data", 0, False
    if len(vals) == 1:
        return "insufficient", vals[0], False
    last = vals[-1]
    diffs = [vals[i] - vals[i - 1] for i in range(1, len(vals))]
    m = statistics.mean(vals)
    sd = statistics.stdev(vals)
    if m and abs(sd / m) < 0.10:
        return "flat", last, False
    up = sum(1 for d in diffs if d > 0)
    total = len(diffs)
    if up / total >= 0.6:
        if pct_values and len(pct_values) >= 2:
            avg_pct = statistics.mean(pct_values)
        else:
            return "growing", last, False
        rc = diffs[-min(3, len(diffs)):]
        if len(rc) >= 2:
            a2 = [rc[i] - rc[i - 1] for i in range(1, len(rc))]
            if all(a > 0 for a in a2):
                return "accelerating", avg_pct, True
            if all(a < 0 for a in a2):
                return "decelerating", avg_pct * 0.92, True
        return "growing", avg_pct, True
    dn = sum(1 for d in diffs if d < 0)
    if dn / total >= 0.6:
        return "declining", round(last + statistics.mean(diffs[-min(3, len(diffs)):])), False
    if cv > 0.4 and len(vals) >= 2:
        decay = 0.85
        w = [decay ** (len(vals) - 1 - i) for i in range(len(vals))]
    else:
        w = list(range(1, len(vals) + 1))
    return "volatile", round(sum(v * wt for v, wt in zip(vals, w)) / sum(w)), False


def compute_momentum(qoq):
    rc = [r["qoq_dollar_change"] for r in qoq[-3:]]
    if len(rc) < 2:
        return "neutral", 1.0
    d = [rc[i] - rc[i - 1] for i in range(1, len(rc))]
    if all(x > 0 for x in d):
        return "accelerating", 1.03
    if all(x < 0 for x in d):
        return "decelerating", 0.97
    return "stable", 1.0


MAX_FWD_DATE = datetime(2028, 1, 31)  # Show up to FY ending closest to Dec 2027


def quarters_to_cutoff(last_period):
    """Count how many quarters from last_period to MAX_FWD_DATE."""
    n = 0
    per = last_period
    while True:
        per = next_period(per, 1)
        if datetime.strptime(per, "%Y-%m-%d") > MAX_FWD_DATE:
            break
        n += 1
    return max(n, 4)


def _stl_project_export(actuals, n_forward):
    """STL with excess-decay dampening — mirrors agent.py."""
    try:
        from statsmodels.tsa.seasonal import STL
        import numpy as np
    except ImportError:
        return None
    revenues = [a["revenue"] for a in actuals]
    if len(revenues) < 12:
        return None
    # Regime truncation: accelerating companies use last 12Q only
    pct_8 = [(revenues[i]-revenues[i-1])/revenues[i-1]
             for i in range(max(1,len(revenues)-8), len(revenues)) if revenues[i-1]>0]
    pct_4 = pct_8[-4:] if len(pct_8)>=4 else pct_8
    if pct_4 and pct_8 and (np.mean(pct_4)-np.mean(pct_8))>0.01 and len(revenues)>12:
        revenues = revenues[-12:]
    series = np.array(revenues, dtype=float)
    stl = STL(series, period=4, robust=True)
    result = stl.fit()
    trend = result.trend
    seasonal_comp = result.seasonal
    trend_diffs = [trend[i] - trend[i - 1] for i in range(max(1, len(trend) - 4), len(trend))]
    avg_inc = float(np.mean(trend_diffs))
    last_trend = float(trend[-1])
    last_seasonal = {}
    for i in range(len(seasonal_comp) - 1, max(len(seasonal_comp) - 5, -1), -1):
        pos = i % 4
        if pos not in last_seasonal:
            last_seasonal[pos] = float(seasonal_comp[i])
    pct_8q = []
    for i in range(max(1, len(revenues) - 8), len(revenues)):
        if revenues[i - 1] > 0:
            pct_8q.append((revenues[i] - revenues[i - 1]) / revenues[i - 1])
    pct_4q = pct_8q[-4:] if len(pct_8q) >= 4 else pct_8q
    avg_8q = float(np.mean(pct_8q)) if pct_8q else 0
    avg_4q = float(np.mean(pct_4q)) if pct_4q else 0
    long_run_pct = avg_4q if (avg_4q - avg_8q) > 0.01 else avg_8q
    EXCESS_DECAY = 0.85
    forward = {}
    last_pos = (len(revenues) - 1) % 4
    prev_rev = float(revenues[-1])
    for i in range(1, n_forward + 1):
        raw_trend = last_trend + avg_inc * i
        pos = (last_pos + i) % 4
        raw_rev = raw_trend + last_seasonal.get(pos, 0)
        stl_pct = (raw_rev - prev_rev) / prev_rev if prev_rev > 0 else 0
        excess = stl_pct - long_run_pct
        dampened_pct = long_run_pct + excess * (EXCESS_DECAY ** (i - 1))
        proj_rev = prev_rev * (1 + dampened_pct)
        forward[i] = round(proj_rev)
        prev_rev = proj_rev
    return forward


def extrapolate(actuals, qoq, estimates, beat_cadence, seasonal, n=None):
    """Q+1: beat-adjusted. Q+2+: STL decomposition (fallback: % QoQ decision tree)."""
    if n is None:
        n = quarters_to_cutoff(actuals[-1]["period"])

    by_q_dollar = defaultdict(list)
    by_q_pct = defaultdict(list)
    for r in qoq:
        by_q_dollar[r["quarter"]].append(r["qoq_dollar_change"])
        by_q_pct[r["quarter"]].append(r["qoq_pct_change"] / 100)

    sf = {}
    for q in ("Q1", "Q2", "Q3", "Q4"):
        all_v = by_q_dollar.get(q, [])
        all_p = by_q_pct.get(q, [])
        if not all_v:
            continue
        v = all_v[-8:]
        pv = all_p[-8:]
        cv = seasonal[q]["cv"] if q in seasonal else 0
        t, p, is_pct = classify_seasonal_trend(v, cv=cv, pct_values=pv)
        sf[q] = {"trend": t, "projected_qoq": p, "is_pct_rate": is_pct, "cv": cv}
    ml, mf = compute_momentum(qoq)

    # STL decomposition for Q+2+
    stl_fwd = _stl_project_export(actuals, n)
    use_stl = stl_fwd is not None

    last_period = actuals[-1]["period"]
    actual_yqs = set()
    for a in actuals:
        ad = datetime.strptime(a["period"], "%Y-%m-%d")
        actual_yqs.add((ad.year, quarter_from_date(a["period"])))
    beat_pct = (beat_cadence["selected_beat_pct"] / 100) if beat_cadence else 0

    est_by_yq = {}
    for e in estimates:
        if e["estimated_revenue"]:
            ed = datetime.strptime(e["period"], "%Y-%m-%d")
            yq = (ed.year, quarter_from_date(e["period"]))
            if yq not in actual_yqs:
                est_by_yq[yq] = e["estimated_revenue"]

    proj = []
    prev = actuals[-1]["revenue"]
    for i in range(1, n + 1):
        per = next_period(last_period, i)
        pq = quarter_from_date(per)
        py = datetime.strptime(per, "%Y-%m-%d").year
        qk = f"Q{pq}"
        con = est_by_yq.get((py, pq))

        if i == 1 and con and beat_cadence:
            rev = round(con * (1 + beat_pct))
            method = "beat_adjusted"
        elif use_stl and i >= 2:
            stl_rev = stl_fwd[i]
            stl_qoq = stl_rev - prev
            rev = prev + stl_qoq
            method = "stl_decomposition"
        else:
            if qk in sf:
                sf_e = sf[qk]
                if sf_e["is_pct_rate"]:
                    bqoq = round(prev * sf_e["projected_qoq"])
                else:
                    bqoq = sf_e["projected_qoq"]
            else:
                bqoq = 0
            aqoq = round(bqoq * mf)
            rev = prev + aqoq
            method = "qoq_extrapolation"
        var_pct = round((rev - con) / con * 100, 2) if con else None
        proj.append({"period": per, "quarter": qk, "projected_revenue": rev,
                      "projected_qoq": rev - prev, "method": method,
                      "consensus": con, "variance_pct": var_pct,
                      "seasonal_trend": sf.get(qk, {}).get("trend", "no_data") if method != "beat_adjusted" else "beat_adjusted",
                      "momentum": ml})
        prev = rev
    return proj, sf, ml, mf


def consensus_comparison(actuals, proj, estimates):
    cy, ny = datetime.now().year, datetime.now().year + 1
    nqc = None
    if proj and proj[0]["consensus"]:
        p = proj[0]
        nqc = {"period": p["period"], "projection": p["projected_revenue"],
               "consensus": p["consensus"],
               "diff_dollars": p["projected_revenue"] - p["consensus"],
               "diff_pct": p["variance_pct"],
               "signal": "BEAT" if (p["variance_pct"] or 0) > 0 else "MISS"}

    def fy(yr):
        at = sum(a["revenue"] for a in actuals if datetime.strptime(a["period"], "%Y-%m-%d").year == yr)
        pt = sum(p["projected_revenue"] for p in proj if datetime.strptime(p["period"], "%Y-%m-%d").year == yr)
        et = sum(e["estimated_revenue"] for e in estimates if datetime.strptime(e["period"], "%Y-%m-%d").year == yr)
        if not et:
            return None
        d = at + pt - et
        return {"fiscal_year": yr, "our_total": at + pt, "consensus_total": et,
                "diff_dollars": d, "diff_pct": round(d / et * 100, 2),
                "signal": "BEAT" if d > 0 else "MISS"}
    return nqc, fy(cy), fy(ny)


def compute_beat_cadence(ticker):
    url = f"https://financialmodelingprep.com/stable/earnings?symbol={ticker}&apikey={os.getenv('FMP_API_KEY')}"
    data = requests.get(url).json()
    if not isinstance(data, list):
        return None
    beats = []
    for e in data:
        ra, re = e.get("revenueActual"), e.get("revenueEstimated")
        if ra and re and re > 0:
            beats.append({"date": e.get("date", ""), "actual": ra, "estimate": re,
                          "beat_pct": round((ra - re) / re * 100, 2)})
    if len(beats) < 2:
        return None
    l4 = [b["beat_pct"] for b in beats[:4]]
    l8 = [b["beat_pct"] for b in beats[:8]]
    a4, a8 = statistics.mean(l4), statistics.mean(l8)
    s4 = statistics.stdev(l4) if len(l4) > 1 else float("inf")
    s8 = statistics.stdev(l8) if len(l8) > 1 else float("inf")
    sel, win = (a4, "4Q") if s4 <= s8 else (a8, "8Q")
    return {"avg_beat_4q": round(a4, 2), "avg_beat_8q": round(a8, 2),
            "std_4q": round(s4, 2) if s4 != float("inf") else 0,
            "std_8q": round(s8, 2) if s8 != float("inf") else 0,
            "selected_beat_pct": round(sel, 2), "selected_window": win,
            "is_changing": abs(a4 - a8) > 1.5, "recent_beats": beats[:8]}


def build_guide_inference(projections, beat_cadence):
    """Beat-cadence driven guide signal (best directional accuracy)."""
    if not beat_cadence or len(projections) < 2:
        return None
    bp = beat_cadence["selected_beat_pct"] / 100
    q2 = projections[1]
    con = q2.get("consensus")
    if not con:
        return None
    ba = round(con * (1 + bp))  # beat-adjusted actual
    ig = round(ba / (1 + bp))   # implied guide
    gd = round(ba - con)
    gp = round((ba - con) / con * 100, 2)
    gs = "GUIDE ABOVE" if gp > 2 else ("GUIDE BELOW" if gp < -2 else "GUIDE IN-LINE")
    return {"period": q2["period"], "quarter": q2["quarter"],
            "projected_actual": q2["projected_revenue"],
            "beat_adjusted_actual": ba,
            "implied_guide": ig,
            "consensus": con, "gap_dollars": gd, "gap_pct": gp, "signal": gs,
            "beat_cadence": beat_cadence}


def flag_anomalies(qoq, seasonal):
    flagged = []
    for row in qoq:
        q = row["quarter"]
        if q in seasonal and seasonal[q]["std_qoq_change"] > 0:
            dev = abs(row["qoq_dollar_change"] - seasonal[q]["avg_qoq_change"]) / seasonal[q]["std_qoq_change"]
            if dev > 1.5:
                flagged.append({"period": row["period"], "quarter": q,
                                "actual_qoq_change": row["qoq_dollar_change"],
                                "seasonal_avg": seasonal[q]["avg_qoq_change"],
                                "std_deviations": round(dev, 2)})
    return flagged


def supplement_estimates_from_earnings(ticker, estimates, actuals):
    """Pull near-term estimates from FMP /earnings that /analyst-estimates misses.

    FMP's /analyst-estimates?period=quarter often skips the next 1-2 quarters,
    but /earnings has the pre-earnings consensus for the upcoming report.
    """
    url = f"https://financialmodelingprep.com/stable/earnings?symbol={ticker}&apikey={os.getenv('FMP_API_KEY')}"
    data = requests.get(url).json()
    if not isinstance(data, list):
        return estimates

    actual_periods = {a["period"] for a in actuals}
    est_periods = {e["period"] for e in estimates}
    last_period = actuals[-1]["period"]

    supplemented = list(estimates)
    for e in data:
        if e.get("revenueActual") is not None:
            continue  # already reported
        re = e.get("revenueEstimated")
        if not re or re <= 0:
            continue
        earnings_date = e.get("date", "")
        if not earnings_date:
            continue
        # Map earnings date to the fiscal quarter-end it covers:
        # earnings happen 1-2 months after quarter end
        earn_d = datetime.strptime(earnings_date, "%Y-%m-%d")
        for steps in range(1, 5):
            per = next_period(last_period, steps)
            per_d = datetime.strptime(per, "%Y-%m-%d")
            gap = (earn_d - per_d).days
            if 0 < gap < 90 and per not in actual_periods and per not in est_periods:
                supplemented.append({"period": per, "estimated_revenue": re})
                est_periods.add(per)
                break

    return supplemented


# ── build all ticker data ─────────────────────────────────────────────────

def build_all():
    out = {}
    for tk in TICKERS:
        print(f"  {tk}...")
        act, est = get_db_data(tk)
        if not act:
            continue
        est = supplement_estimates_from_earnings(tk, est, act)
        qoq = compute_qoq(act)
        sea = compute_seasonality(qoq)
        anom = flag_anomalies(qoq, sea)
        bc = compute_beat_cadence(tk)
        proj, sf, ml, mf = extrapolate(act, qoq, est, bc, sea)
        gi = build_guide_inference(proj, bc)
        nq, cfy, nfy = consensus_comparison(act, proj, est)
        ta_map = get_transcript_analyses(tk)
        pec_map = get_pre_earnings_consensus(tk)
        for a in anom:
            ta = ta_map.get(a["period"])
            if ta:
                a["transcript_analysis"] = ta
        out[tk] = {
            "actuals": act, "estimates": est, "qoq": qoq,
            "projections": proj, "seasonal_forecasts": sf,
            "momentum_label": ml, "momentum_factor": mf,
            "beat_cadence": bc, "guide_inference": gi,
            "anomalies": anom, "transcript_analyses": ta_map,
            "pre_earnings_consensus": pec_map,
            "consensus": {"next_quarter": nq, "current_fy": cfy, "next_fy": nfy},
        }
    return out


# ── styles ────────────────────────────────────────────────────────────────

NAVY = "1A1F2E"
BLUE = "0000FF"
GREEN_TXT = "006100"
RED_TXT = "9C0006"
BLACK = "000000"
WHITE = "FFFFFF"

# Section headers: bold, not italic
F_TITLE = Font(name="Times New Roman", size=12, bold=True)
F_SUBTITLE = Font(name="Times New Roman", size=9, italic=True)
F_BOLD = Font(name="Times New Roman", size=9, bold=True)

# Sub-rows: italic
F_IT = Font(name="Times New Roman", size=9, italic=True)
F_IT_BLUE = Font(name="Times New Roman", size=9, italic=True, color=BLUE)
F_IT_GREEN = Font(name="Times New Roman", size=9, italic=True, color=GREEN_TXT)
F_IT_RED = Font(name="Times New Roman", size=9, italic=True, color=RED_TXT)

# Header block values
F_BLUE = Font(name="Times New Roman", size=9, color=BLUE)
F_DATA = Font(name="Times New Roman", size=9)
F_PERIOD = Font(name="Times New Roman", size=9, bold=True)

# Summary-only
F_SEC = Font(name="Times New Roman", size=10, bold=True, color=WHITE)
FILL_SEC = PatternFill("solid", fgColor=NAVY)
FILL_G = PatternFill("solid", fgColor="C6EFCE")
FILL_R = PatternFill("solid", fgColor="FFC7CE")
FILL_Y = PatternFill("solid", fgColor="FFEB9C")

N_REV = '#,##0.0_);(#,##0.0);"-"'
N_PCT = '0.0%_);(0.0%);"-"'
N_DLR = '#,##0.0_);(#,##0.0);"-"'

A_C = Alignment(horizontal="center")
A_R = Alignment(horizontal="right")
A_L = Alignment(horizontal="left")

FY_BORDER = Side(style="thin")


def cr(row, col):
    return f"{get_column_letter(col)}{row}"


def prev_q(y, q):
    return (y - 1, 4) if q == 1 else (y, q - 1)


# ── column layout builder ────────────────────────────────────────────────

def build_columns(actuals, projections, estimates=None):
    """Build wide time-series column list: quarters + FY totals.

    Truncates at MAX_FWD_DATE. All forward quarters are covered by extrapolation
    (no consensus-only columns).
    """
    entries = {}
    for a in actuals:
        d = datetime.strptime(a["period"], "%Y-%m-%d")
        q = quarter_from_date(a["period"])
        entries[(d.year, q)] = {"rev": a["revenue"], "est": False, "per": a["period"],
                                "method": None, "proj_qoq": None}
    for p in projections:
        d = datetime.strptime(p["period"], "%Y-%m-%d")
        if d > MAX_FWD_DATE:
            continue  # truncate beyond cutoff
        q = quarter_from_date(p["period"])
        if (d.year, q) not in entries:
            entries[(d.year, q)] = {"rev": p["projected_revenue"], "est": True,
                                    "per": p["period"], "method": p["method"],
                                    "proj_qoq": p["projected_qoq"]}

    est_years = {y for (y, _), e in entries.items() if e["est"]}
    first_est_y = min(est_years) if est_years else max(y for y, _ in entries)
    show_from = first_est_y - 2

    cols = []
    ci = DATA_COL
    fy_groups = {}

    for yr in sorted(set(y for y, _ in entries if y >= show_from)):
        yr_keys = sorted((y, q) for y, q in entries if y == yr and y >= show_from)
        first_ci_for_yr = ci
        for (y, q) in yr_keys:
            e = entries[(y, q)]
            yy = str(y)[-2:]
            cols.append({"t": "q", "lbl": f"Q{q}-{yy}{'E' if e['est'] else ''}",
                         "y": y, "q": q, "rev": e["rev"], "est": e["est"],
                         "per": e["per"], "ci": ci,
                         "method": e.get("method"), "proj_qoq": e.get("proj_qoq")})
            ci += 1
        yr_qs = [q for (y2, q) in yr_keys]
        if set(yr_qs) == {1, 2, 3, 4}:
            has_e = any(entries[(yr, q)]["est"] for q in (1, 2, 3, 4))
            yy = str(yr)[-2:]
            fq = [c["ci"] for c in cols if c["t"] == "q" and c["y"] == yr]
            cols.append({"t": "fy", "lbl": f"FY{yy}{'E' if has_e else ''}",
                         "y": yr, "q": 0, "est": has_e, "ci": ci, "fqc": fq})
            fy_groups[yr] = {"first_ci": first_ci_for_yr, "fy_ci": ci}
            ci += 1
        else:
            fy_groups[yr] = {"first_ci": first_ci_for_yr, "fy_ci": None}

    qmap = {(c["y"], c["q"]): c["ci"] for c in cols if c["t"] == "q"}
    fymap = {c["y"]: c["ci"] for c in cols if c["t"] == "fy"}
    return cols, qmap, fymap, fy_groups


# ── ticker sheet ──────────────────────────────────────────────────────────

def build_ticker_sheet(wb, ticker, data):
    ws = wb.create_sheet(title=ticker)
    ws.sheet_view.showGridLines = False

    bc = data["beat_cadence"]
    gi = data["guide_inference"]
    proj = data["projections"]
    anomalies = data["anomalies"]
    ta_map = data.get("transcript_analyses", {})

    # Consensus lookup by (year, quarter) — matches regardless of day-of-month
    est_yq_lookup = {}
    for e in data["estimates"]:
        if e["estimated_revenue"]:
            ed = datetime.strptime(e["period"], "%Y-%m-%d")
            est_yq_lookup[(ed.year, quarter_from_date(e["period"]))] = e["estimated_revenue"]
    # Projection lookup
    proj_map = {p["period"]: p for p in proj}
    # Anomaly periods
    anomaly_periods = {a["period"] for a in anomalies}

    cols, qmap, fymap, fy_groups = build_columns(data["actuals"], proj, data["estimates"])
    last_ci = cols[-1]["ci"] if cols else DATA_COL
    beat_ref = "$C$3"

    # ── Row assignments ──
    R_PER = 7   # period headers
    # row 8 empty
    R_REV = 9   # Total Revenue
    R_YP = 10   # % YoY
    R_QP = 11   # % QoQ
    R_YD = 12   # $ YoY
    R_QD = 13   # $ QoQ (blue driver for forward)
    # row 14 empty
    R_CON = 15  # Consensus Total Revenue
    R_CYP = 16  # Consensus % YoY
    R_CQP = 17  # Consensus % QoQ
    R_CYD = 18  # Consensus $ YoY
    R_CQD = 19  # Consensus $ QoQ
    # row 20 empty
    R_VAR = 21  # % Variance vs Consensus
    # row 22 empty
    R_GD = 23   # Implied Q+2 Guide
    R_GYP = 24  # Guide % YoY
    R_GQP = 25  # Guide % QoQ
    R_GYD = 26  # Guide $ YoY
    R_GQD = 27  # Guide $ QoQ
    # row 28 empty
    R_GV = 29   # % Variance Guide vs Consensus
    # row 30 empty
    R_AVC = 31  # "Actuals vs Consensus" section header
    R_AREV = 32 # Actual Revenue
    R_ACON = 33 # Pre-earnings Consensus Revenue
    R_ABD = 34  # Beat / Miss ($)
    R_ABP = 35  # Beat / Miss (%)
    # row 36 empty
    R_T4 = 37   # Trailing 4Q Avg Beat
    R_T8 = 38   # Trailing 8Q Avg Beat
    R_SEL = 39  # Selected Beat Cadence

    R_BOT = R_SEL

    # ── Column widths ──
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 14
    for c in cols:
        ws.column_dimensions[get_column_letter(c["ci"])].width = 11.5 if c["t"] == "fy" else 10.5

    # ════════════════════════════════════════════════════════════════════════
    #  HEADER BLOCK (rows 1-5)
    # ════════════════════════════════════════════════════════════════════════

    ws.cell(row=1, column=2, value=f"{ticker}: KIN Base Case Operating Model").font = F_TITLE
    ws.cell(row=2, column=2, value="USD in Millions Unless Stated Otherwise").font = F_SUBTITLE

    ws.cell(row=3, column=2, value="Beat Cadence (%)").font = F_DATA
    if bc:
        c = ws.cell(row=3, column=3, value=bc["selected_beat_pct"] / 100)
        c.font = F_BLUE
        c.number_format = "0.00%"
    ws.cell(row=4, column=2, value="Beat Window").font = F_DATA
    if bc:
        ws.cell(row=4, column=3, value=bc["selected_window"]).font = F_BLUE
    ws.cell(row=5, column=2, value="Momentum").font = F_DATA
    ws.cell(row=5, column=3, value=data["momentum_label"].upper()).font = F_BLUE

    # ════════════════════════════════════════════════════════════════════════
    #  PERIOD HEADERS (row 7)
    # ════════════════════════════════════════════════════════════════════════

    for c in cols:
        cell = ws.cell(row=R_PER, column=c["ci"], value=c["lbl"])
        cell.font = F_PERIOD
        cell.alignment = A_C

    # ════════════════════════════════════════════════════════════════════════
    #  FISCAL YEAR BORDERS (left of Q1, right of FY)
    # ════════════════════════════════════════════════════════════════════════

    for yr, grp in fy_groups.items():
        for r in range(R_PER, R_BOT + 1):
            # Left border on first quarter of the year
            cell = ws.cell(row=r, column=grp["first_ci"])
            cell.border = Border(left=FY_BORDER, right=cell.border.right,
                                 top=cell.border.top, bottom=cell.border.bottom)
            # FY column: both left AND right borders
            if grp["fy_ci"]:
                cell = ws.cell(row=r, column=grp["fy_ci"])
                cell.border = Border(left=FY_BORDER, right=FY_BORDER,
                                     top=cell.border.top, bottom=cell.border.bottom)

    # ════════════════════════════════════════════════════════════════════════
    #  ROW LABELS
    # ════════════════════════════════════════════════════════════════════════

    ws.cell(row=R_REV, column=2, value="Total Revenue").font = F_BOLD
    ws.cell(row=R_YP, column=2, value="   % YoY").font = F_IT
    ws.cell(row=R_QP, column=2, value="   % QoQ").font = F_IT
    ws.cell(row=R_YD, column=2, value="   $ YoY").font = F_IT
    ws.cell(row=R_QD, column=2, value="   $ QoQ").font = F_IT

    ws.cell(row=R_CON, column=2, value="Consensus Total Revenue").font = F_BOLD
    ws.cell(row=R_CYP, column=2, value="   % YoY").font = F_IT
    ws.cell(row=R_CQP, column=2, value="   % QoQ").font = F_IT
    ws.cell(row=R_CYD, column=2, value="   $ YoY").font = F_IT
    ws.cell(row=R_CQD, column=2, value="   $ QoQ").font = F_IT

    ws.cell(row=R_VAR, column=2, value="   % Variance vs Consensus").font = F_IT

    ws.cell(row=R_GD, column=2, value="Implied Q+2 Guide").font = F_BOLD
    ws.cell(row=R_GYP, column=2, value="   % YoY").font = F_IT
    ws.cell(row=R_GQP, column=2, value="   % QoQ").font = F_IT
    ws.cell(row=R_GYD, column=2, value="   $ YoY").font = F_IT
    ws.cell(row=R_GQD, column=2, value="   $ QoQ").font = F_IT

    ws.cell(row=R_GV, column=2, value="   % Variance Guide vs Consensus").font = F_IT

    # ════════════════════════════════════════════════════════════════════════
    #  Q+2 column index for guide section
    # ════════════════════════════════════════════════════════════════════════

    gi_ci = None
    if gi:
        gd = datetime.strptime(gi["period"], "%Y-%m-%d")
        gi_ci = qmap.get((gd.year, quarter_from_date(gi["period"])))

    # Q+1 column index (for guide QoQ reference)
    q1_ci = None
    if proj:
        q1d = datetime.strptime(proj[0]["period"], "%Y-%m-%d")
        q1_ci = qmap.get((q1d.year, quarter_from_date(proj[0]["period"])))

    # ════════════════════════════════════════════════════════════════════════
    #  TOTAL REVENUE ROW
    #  Historical: hardcoded black | Forward: formula = prev Q + $ QoQ
    # ════════════════════════════════════════════════════════════════════════

    for col in cols:
        ci = col["ci"]
        if col["t"] == "q":
            if not col["est"]:
                # HISTORICAL: hardcoded revenue — black text
                cell = ws.cell(row=R_REV, column=ci, value=col["rev"] / 1e6)
                cell.font = F_BOLD
                cell.number_format = N_REV
                cell.alignment = A_R

                # Cell comment for anomalous quarters with transcript analysis
                if col["per"] in anomaly_periods and col["per"] in ta_map:
                    raw = ta_map[col["per"]]
                    lines = raw.strip().split("\n")
                    short = "\n".join(lines[:4])
                    if len(short) > 300:
                        short = short[:297] + "..."
                    cell.comment = Comment(short, "KIN Model")
            else:
                # FORWARD: formula = prior Q revenue + $ QoQ driver
                py, pq = prev_q(col["y"], col["q"])
                prev_ci = qmap.get((py, pq))
                if prev_ci:
                    cell = ws.cell(row=R_REV, column=ci,
                                   value=f"={cr(R_REV, prev_ci)}+{cr(R_QD, ci)}")
                elif col["rev"] is not None:
                    cell = ws.cell(row=R_REV, column=ci, value=col["rev"] / 1e6)
                else:
                    continue
                cell.font = F_BOLD
                cell.number_format = N_REV
                cell.alignment = A_R

                # Cell comment: beat cadence math for Q+1
                pm = proj_map.get(col["per"])
                if pm and pm["method"] == "beat_adjusted" and pm["consensus"] and bc:
                    cell.comment = Comment(
                        f"Q+1 Beat-Adjusted:\n"
                        f"Consensus: ${pm['consensus']/1e6:,.1f}M\n"
                        f"x (1 + {bc['selected_beat_pct']:.2f}%)\n"
                        f"= ${pm['projected_revenue']/1e6:,.1f}M implied",
                        "KIN Model")

        elif col["t"] == "fy":
            # FY: SUM of Q columns
            fq = col["fqc"]
            cell = ws.cell(row=R_REV, column=ci,
                           value=f"=SUM({cr(R_REV, fq[0])}:{cr(R_REV, fq[-1])})")
            cell.font = F_BOLD
            cell.number_format = N_REV
            cell.alignment = A_R

    # ════════════════════════════════════════════════════════════════════════
    #  $ QoQ ROW
    #  Historical: formula | Forward: hardcoded blue (the assumption input)
    # ════════════════════════════════════════════════════════════════════════

    for col in cols:
        if col["t"] != "q":
            continue
        ci = col["ci"]
        py, pq = prev_q(col["y"], col["q"])
        prev_ci = qmap.get((py, pq))
        if not prev_ci:
            continue

        if not col["est"]:
            # HISTORICAL: formula
            cell = ws.cell(row=R_QD, column=ci,
                           value=f"={cr(R_REV, ci)}-{cr(R_REV, prev_ci)}")
            cell.font = F_IT
        else:
            # FORWARD: hardcoded blue — this is the extrapolation driver
            if col["proj_qoq"] is not None:
                cell = ws.cell(row=R_QD, column=ci,
                               value=col["proj_qoq"] / 1e6)
                cell.font = F_IT_BLUE
            else:
                # Consensus-only column beyond projection window — skip
                continue
        cell.number_format = N_DLR
        cell.alignment = A_R

    # ════════════════════════════════════════════════════════════════════════
    #  % YoY, % QoQ, $ YoY — all formula-driven, italic
    # ════════════════════════════════════════════════════════════════════════

    for col in cols:
        ci = col["ci"]

        # % YoY
        if col["t"] == "q":
            yr_ci = qmap.get((col["y"] - 1, col["q"]))
            if yr_ci:
                cell = ws.cell(row=R_YP, column=ci,
                               value=f'=IFERROR(({cr(R_REV, ci)}-{cr(R_REV, yr_ci)})/{cr(R_REV, yr_ci)},"-")')
                cell.font = F_IT
                cell.number_format = N_PCT
                cell.alignment = A_R
        elif col["t"] == "fy":
            pfy = fymap.get(col["y"] - 1)
            if pfy:
                cell = ws.cell(row=R_YP, column=ci,
                               value=f'=IFERROR(({cr(R_REV, ci)}-{cr(R_REV, pfy)})/{cr(R_REV, pfy)},"-")')
                cell.font = F_IT
                cell.number_format = N_PCT
                cell.alignment = A_R

        # % QoQ (quarters only)
        if col["t"] == "q":
            py, pq = prev_q(col["y"], col["q"])
            prev_ci = qmap.get((py, pq))
            if prev_ci:
                cell = ws.cell(row=R_QP, column=ci,
                               value=f'=IFERROR(({cr(R_REV, ci)}-{cr(R_REV, prev_ci)})/{cr(R_REV, prev_ci)},"-")')
                cell.font = F_IT
                cell.number_format = N_PCT
                cell.alignment = A_R

        # $ YoY
        if col["t"] == "q":
            yr_ci = qmap.get((col["y"] - 1, col["q"]))
            if yr_ci:
                cell = ws.cell(row=R_YD, column=ci,
                               value=f"={cr(R_REV, ci)}-{cr(R_REV, yr_ci)}")
                cell.font = F_IT
                cell.number_format = N_DLR
                cell.alignment = A_R
        elif col["t"] == "fy":
            pfy = fymap.get(col["y"] - 1)
            if pfy:
                cell = ws.cell(row=R_YD, column=ci,
                               value=f"={cr(R_REV, ci)}-{cr(R_REV, pfy)}")
                cell.font = F_IT
                cell.number_format = N_DLR
                cell.alignment = A_R

    # ════════════════════════════════════════════════════════════════════════
    #  CONSENSUS TOTAL REVENUE ROW
    #  Reported quarters: blank | Unreported with DB estimate: hardcoded
    # ════════════════════════════════════════════════════════════════════════

    actual_periods = {a["period"] for a in data["actuals"]}
    # Track which Q columns got a consensus value (for FY SUM logic)
    con_cols = set()

    for col in cols:
        ci = col["ci"]
        if col["t"] == "q":
            if col["per"] in actual_periods:
                # REPORTED: blank — consensus not meaningful for reported quarters
                continue
            # UNREPORTED: only show if consensus exists in DB
            con_val = est_yq_lookup.get((col["y"], col["q"]))
            pm = proj_map.get(col["per"])
            if con_val:
                cell = ws.cell(row=R_CON, column=ci, value=con_val / 1e6)
            elif pm and pm["consensus"]:
                cell = ws.cell(row=R_CON, column=ci, value=pm["consensus"] / 1e6)
            else:
                continue
            cell.font = F_BOLD
            cell.number_format = N_REV
            cell.alignment = A_R
            con_cols.add(ci)
        elif col["t"] == "fy":
            # Only show FY consensus if all 4 Q columns have consensus
            fq = col["fqc"]
            if all(fci in con_cols for fci in fq):
                cell = ws.cell(row=R_CON, column=ci,
                               value=f"=SUM({cr(R_CON, fq[0])}:{cr(R_CON, fq[-1])})")
                cell.font = F_BOLD
                cell.number_format = N_REV
                cell.alignment = A_R
                con_cols.add(ci)

    # ── Consensus memo lines — only for columns that have consensus data ──

    for col in cols:
        ci = col["ci"]
        if ci not in con_cols:
            continue

        # Consensus % YoY
        if col["t"] == "q":
            yr_ci = qmap.get((col["y"] - 1, col["q"]))
            if yr_ci and yr_ci in con_cols:
                cell = ws.cell(row=R_CYP, column=ci,
                               value=f'=IFERROR(({cr(R_CON, ci)}-{cr(R_CON, yr_ci)})/{cr(R_CON, yr_ci)},"-")')
                cell.font = F_IT
                cell.number_format = N_PCT
                cell.alignment = A_R
        elif col["t"] == "fy":
            pfy = fymap.get(col["y"] - 1)
            if pfy and pfy in con_cols:
                cell = ws.cell(row=R_CYP, column=ci,
                               value=f'=IFERROR(({cr(R_CON, ci)}-{cr(R_CON, pfy)})/{cr(R_CON, pfy)},"-")')
                cell.font = F_IT
                cell.number_format = N_PCT
                cell.alignment = A_R

        # Consensus % QoQ (quarters only)
        if col["t"] == "q":
            py, pq = prev_q(col["y"], col["q"])
            prev_ci = qmap.get((py, pq))
            if prev_ci and prev_ci in con_cols:
                cell = ws.cell(row=R_CQP, column=ci,
                               value=f'=IFERROR(({cr(R_CON, ci)}-{cr(R_CON, prev_ci)})/{cr(R_CON, prev_ci)},"-")')
                cell.font = F_IT
                cell.number_format = N_PCT
                cell.alignment = A_R

        # Consensus $ YoY
        if col["t"] == "q":
            yr_ci = qmap.get((col["y"] - 1, col["q"]))
            if yr_ci and yr_ci in con_cols:
                cell = ws.cell(row=R_CYD, column=ci,
                               value=f"={cr(R_CON, ci)}-{cr(R_CON, yr_ci)}")
                cell.font = F_IT
                cell.number_format = N_DLR
                cell.alignment = A_R
        elif col["t"] == "fy":
            pfy = fymap.get(col["y"] - 1)
            if pfy and pfy in con_cols:
                cell = ws.cell(row=R_CYD, column=ci,
                               value=f"={cr(R_CON, ci)}-{cr(R_CON, pfy)}")
                cell.font = F_IT
                cell.number_format = N_DLR
                cell.alignment = A_R

        # Consensus $ QoQ (quarters only)
        if col["t"] == "q":
            py, pq = prev_q(col["y"], col["q"])
            prev_ci = qmap.get((py, pq))
            if prev_ci and prev_ci in con_cols:
                cell = ws.cell(row=R_CQD, column=ci,
                               value=f"={cr(R_CON, ci)}-{cr(R_CON, prev_ci)}")
                cell.font = F_IT
                cell.number_format = N_DLR
                cell.alignment = A_R

    # ════════════════════════════════════════════════════════════════════════
    #  % VARIANCE VS CONSENSUS — forward quarters only, formula, green/red
    # ════════════════════════════════════════════════════════════════════════

    for col in cols:
        if col["t"] != "q" or not col["est"]:
            continue
        ci = col["ci"]
        # Only if consensus exists for this column
        con_val = est_yq_lookup.get((col["y"], col["q"]))
        pm = proj_map.get(col["per"])
        has_con = con_val or (pm and pm["consensus"])
        if not has_con:
            continue

        f = f'=IFERROR(({cr(R_REV, ci)}-{cr(R_CON, ci)})/{cr(R_CON, ci)},"-")'
        cell = ws.cell(row=R_VAR, column=ci, value=f)
        cell.font = F_IT
        cell.number_format = N_PCT
        cell.alignment = A_R

    # FY variance
    for col in cols:
        if col["t"] != "fy" or not col["est"]:
            continue
        ci = col["ci"]
        f = f'=IFERROR(({cr(R_REV, ci)}-{cr(R_CON, ci)})/{cr(R_CON, ci)},"-")'
        cell = ws.cell(row=R_VAR, column=ci, value=f)
        cell.font = F_IT
        cell.number_format = N_PCT
        cell.alignment = A_R

    # ════════════════════════════════════════════════════════════════════════
    #  IMPLIED Q+2 GUIDE — one column, formula-driven
    # ════════════════════════════════════════════════════════════════════════

    if gi and gi_ci and bc:
        gd = datetime.strptime(gi["period"], "%Y-%m-%d")
        gq = quarter_from_date(gi["period"])

        # Implied guide: our STL revenue estimate / (1 + beat %)
        # Infers what management will guide to given our projected actual
        cell = ws.cell(row=R_GD, column=gi_ci,
                       value=f"={cr(R_REV, gi_ci)}/(1+{beat_ref})")
        cell.font = F_BOLD
        cell.number_format = N_REV
        cell.alignment = A_R

        # Guide % YoY
        ya_ci = qmap.get((gd.year - 1, gq))
        if ya_ci:
            cell = ws.cell(row=R_GYP, column=gi_ci,
                           value=f'=IFERROR(({cr(R_GD, gi_ci)}-{cr(R_REV, ya_ci)})/{cr(R_REV, ya_ci)},"-")')
            cell.font = F_IT
            cell.number_format = N_PCT
            cell.alignment = A_R

        # Guide % QoQ (vs Q+1 in revenue row)
        if q1_ci:
            cell = ws.cell(row=R_GQP, column=gi_ci,
                           value=f'=IFERROR(({cr(R_GD, gi_ci)}-{cr(R_REV, q1_ci)})/{cr(R_REV, q1_ci)},"-")')
            cell.font = F_IT
            cell.number_format = N_PCT
            cell.alignment = A_R

        # Guide $ YoY
        if ya_ci:
            cell = ws.cell(row=R_GYD, column=gi_ci,
                           value=f"={cr(R_GD, gi_ci)}-{cr(R_REV, ya_ci)}")
            cell.font = F_IT
            cell.number_format = N_DLR
            cell.alignment = A_R

        # Guide $ QoQ
        if q1_ci:
            cell = ws.cell(row=R_GQD, column=gi_ci,
                           value=f"={cr(R_GD, gi_ci)}-{cr(R_REV, q1_ci)}")
            cell.font = F_IT
            cell.number_format = N_DLR
            cell.alignment = A_R

        # % Variance guide vs consensus — formula, green/red via conditional formatting
        cell = ws.cell(row=R_GV, column=gi_ci,
                       value=f'=IFERROR(({cr(R_GD, gi_ci)}-{cr(R_CON, gi_ci)})/{cr(R_CON, gi_ci)},"-")')
        cell.font = F_IT
        cell.number_format = N_PCT
        cell.alignment = A_R

    # ════════════════════════════════════════════════════════════════════════
    #  ACTUALS VS CONSENSUS — historical quarters only
    # ════════════════════════════════════════════════════════════════════════

    pec_map = data.get("pre_earnings_consensus", {})

    ws.cell(row=R_AVC, column=2, value="Actuals vs Consensus").font = F_BOLD
    ws.cell(row=R_AREV, column=2, value="Actual Revenue").font = F_BOLD
    ws.cell(row=R_ACON, column=2, value="Consensus Revenue").font = F_BOLD
    ws.cell(row=R_ABD, column=2, value="   Beat / Miss ($)").font = F_IT
    ws.cell(row=R_ABP, column=2, value="   Beat / Miss (%)").font = F_IT
    ws.cell(row=R_T4, column=2, value="   Trailing 4Q Avg Beat").font = F_IT
    ws.cell(row=R_T8, column=2, value="   Trailing 8Q Avg Beat").font = F_IT
    ws.cell(row=R_SEL, column=2, value="   Selected Beat Cadence").font = F_IT

    # Selected beat cadence — hardcoded model input (blue)
    if bc:
        c = ws.cell(row=R_SEL, column=3, value=bc["selected_beat_pct"] / 100)
        c.font = F_IT_BLUE
        c.number_format = N_PCT
    # Trailing 4Q/8Q formulas are set after the beat row is populated (see below)

    # Per-column data: actuals, pre-earnings consensus, beat formulas
    beat_cols = []  # track column indices that have beat % data
    for col in cols:
        ci = col["ci"]

        if col["t"] == "q" and not col["est"]:
            per = col["per"]

            # Actual Revenue — hardcoded
            cell = ws.cell(row=R_AREV, column=ci, value=col["rev"] / 1e6)
            cell.font = F_BOLD
            cell.number_format = N_REV
            cell.alignment = A_R

            # Pre-earnings Consensus — hardcoded from pre_earnings_consensus table
            pec = pec_map.get(per)
            if pec:
                cell = ws.cell(row=R_ACON, column=ci, value=pec / 1e6)
                cell.font = F_BOLD
                cell.number_format = N_REV
                cell.alignment = A_R

                # Beat / Miss ($) — formula
                cell = ws.cell(row=R_ABD, column=ci,
                               value=f"={cr(R_AREV, ci)}-{cr(R_ACON, ci)}")
                cell.font = F_IT
                cell.number_format = N_DLR
                cell.alignment = A_R

                # Beat / Miss (%) — formula, green/red
                cell = ws.cell(row=R_ABP, column=ci,
                               value=f'=IFERROR(({cr(R_AREV, ci)}-{cr(R_ACON, ci)})/{cr(R_ACON, ci)},"-")')
                beat_val = (col["rev"] - pec) / pec
                cell.font = F_IT_GREEN if beat_val >= 0 else F_IT_RED
                cell.number_format = N_PCT
                cell.alignment = A_R
                beat_cols.append(ci)

        elif col["t"] == "fy" and not col["est"]:
            # FY actuals sum
            fq = col["fqc"]
            # Only if all Q columns have pre-earnings data
            has_all = all(
                any(c2["ci"] == fci and not c2["est"] and c2["per"] in pec_map
                    for c2 in cols if c2["t"] == "q")
                for fci in fq
            )
            cell = ws.cell(row=R_AREV, column=ci,
                           value=f"=SUM({cr(R_AREV, fq[0])}:{cr(R_AREV, fq[-1])})")
            cell.font = F_BOLD
            cell.number_format = N_REV
            cell.alignment = A_R

            if has_all:
                cell = ws.cell(row=R_ACON, column=ci,
                               value=f"=SUM({cr(R_ACON, fq[0])}:{cr(R_ACON, fq[-1])})")
                cell.font = F_BOLD
                cell.number_format = N_REV
                cell.alignment = A_R

                cell = ws.cell(row=R_ABD, column=ci,
                               value=f"={cr(R_AREV, ci)}-{cr(R_ACON, ci)}")
                cell.font = F_IT
                cell.number_format = N_DLR
                cell.alignment = A_R

                cell = ws.cell(row=R_ABP, column=ci,
                               value=f'=IFERROR(({cr(R_AREV, ci)}-{cr(R_ACON, ci)})/{cr(R_ACON, ci)},"-")')
                cell.font = F_IT
                cell.number_format = N_PCT
                cell.alignment = A_R

    # ── Trailing 4Q/8Q Avg Beat — live AVERAGE formulas from beat % row ──
    if len(beat_cols) >= 4:
        last4 = beat_cols[-4:]
        refs4 = ",".join(cr(R_ABP, c) for c in last4)
        c = ws.cell(row=R_T4, column=3, value=f"=AVERAGE({refs4})")
        c.font = F_IT
        c.number_format = N_PCT
    if len(beat_cols) >= 8:
        last8 = beat_cols[-8:]
        refs8 = ",".join(cr(R_ABP, c) for c in last8)
        c = ws.cell(row=R_T8, column=3, value=f"=AVERAGE({refs8})")
        c.font = F_IT
        c.number_format = N_PCT
    elif len(beat_cols) >= 2:
        # Fewer than 8 quarters: average all available
        refs_all = ",".join(cr(R_ABP, c) for c in beat_cols)
        c = ws.cell(row=R_T8, column=3, value=f"=AVERAGE({refs_all})")
        c.font = F_IT
        c.number_format = N_PCT

    # ── Conditional formatting: green text for positive variance, red for negative ──
    from openpyxl.formatting.rule import FormulaRule
    green_font = Font(name="Times New Roman", size=9, italic=True, color=GREEN_TXT)
    red_font = Font(name="Times New Roman", size=9, italic=True, color=RED_TXT)
    var_range = f"{get_column_letter(DATA_COL)}{R_VAR}:{get_column_letter(last_ci)}{R_VAR}"
    ws.conditional_formatting.add(var_range, FormulaRule(
        formula=[f'{get_column_letter(DATA_COL)}{R_VAR}>0'], font=green_font))
    ws.conditional_formatting.add(var_range, FormulaRule(
        formula=[f'{get_column_letter(DATA_COL)}{R_VAR}<0'], font=red_font))
    gv_range = f"{get_column_letter(DATA_COL)}{R_GV}:{get_column_letter(last_ci)}{R_GV}"
    ws.conditional_formatting.add(gv_range, FormulaRule(
        formula=[f'{get_column_letter(DATA_COL)}{R_GV}>0'], font=green_font))
    ws.conditional_formatting.add(gv_range, FormulaRule(
        formula=[f'{get_column_letter(DATA_COL)}{R_GV}<0'], font=red_font))

    # ── freeze panes ──
    ws.freeze_panes = f"{get_column_letter(DATA_COL)}{R_REV}"


def _total_row_border(ws, row, last_ci):
    """Apply thin top border across a Total Revenue row on the summary sheet."""
    for ci in range(1, last_ci + 1):
        cell = ws.cell(row=row, column=ci)
        cell.border = Border(top=FY_BORDER, left=cell.border.left,
                             right=cell.border.right, bottom=cell.border.bottom)


# ── summary sheet ─────────────────────────────────────────────────────────


def _summary_columns(all_data):
    """Build unified time-series columns across all tickers for the summary sheet.

    Returns (cols, fy_groups) where cols is a list of dicts with t/lbl/y/q/ci/fqc keys,
    and fy_groups maps year -> {first_ci, fy_ci}.
    """
    all_entries = {}
    for tk, d in all_data.items():
        for a in d["actuals"]:
            dt = datetime.strptime(a["period"], "%Y-%m-%d")
            yq = (dt.year, quarter_from_date(a["period"]))
            if yq not in all_entries:
                all_entries[yq] = {"est": False}
        for p in d["projections"]:
            dt = datetime.strptime(p["period"], "%Y-%m-%d")
            if dt > MAX_FWD_DATE:
                continue  # truncate beyond cutoff
            yq = (dt.year, quarter_from_date(p["period"]))
            if yq not in all_entries:
                all_entries[yq] = {"est": True}

    est_years = {y for (y, _), e in all_entries.items() if e["est"]}
    first_est_y = min(est_years) if est_years else max(y for y, _ in all_entries)
    show_from = first_est_y - 2

    cols = []
    ci = 4  # col D = first data column (A=ticker, B=beat%, C=momentum)
    fy_groups = {}

    for yr in sorted(set(y for y, _ in all_entries if y >= show_from)):
        yr_keys = sorted((y, q) for y, q in all_entries if y == yr and y >= show_from)
        first_ci_for_yr = ci
        for (y, q) in yr_keys:
            e = all_entries[(y, q)]
            yy = str(y)[-2:]
            cols.append({"t": "q", "lbl": f"Q{q}-{yy}{'E' if e['est'] else ''}",
                         "y": y, "q": q, "est": e["est"], "ci": ci})
            ci += 1
        yr_qs = [q for (y2, q) in yr_keys]
        if set(yr_qs) == {1, 2, 3, 4}:
            has_e = any(all_entries[(yr, q)]["est"] for q in (1, 2, 3, 4))
            yy = str(yr)[-2:]
            fq = [c["ci"] for c in cols if c["t"] == "q" and c["y"] == yr]
            cols.append({"t": "fy", "lbl": f"FY{yy}{'E' if has_e else ''}",
                         "y": yr, "q": 0, "est": has_e, "ci": ci, "fqc": fq})
            fy_groups[yr] = {"first_ci": first_ci_for_yr, "fy_ci": ci}
            ci += 1
        else:
            fy_groups[yr] = {"first_ci": first_ci_for_yr, "fy_ci": None}

    qmap = {(c["y"], c["q"]): c["ci"] for c in cols if c["t"] == "q"}
    fymap = {c["y"]: c["ci"] for c in cols if c["t"] == "fy"}
    return cols, fy_groups, qmap, fymap


def _ticker_col_for_summary(tk_sheet_name, tk_data, summary_col):
    """Find the column letter on a ticker sheet that matches a summary (year, quarter).

    Ticker sheets start data at DATA_COL=4. We match by (year, quarter).
    Returns the column letter string or None.
    """
    for a in tk_data["actuals"]:
        d = datetime.strptime(a["period"], "%Y-%m-%d")
        if (d.year, quarter_from_date(a["period"])) == (summary_col["y"], summary_col["q"]):
            # Find column index on ticker sheet by counting from DATA_COL
            return None  # we'll compute in the main loop
    return None


def build_summary_sheet(wb, all_data):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    F_MOM_GREEN = Font(name="Times New Roman", size=9, color=GREEN_TXT)
    F_MOM_RED = Font(name="Times New Roman", size=9, color=RED_TXT)

    # Sort tickers by most recent full FY revenue descending
    def fy_revenue(tk):
        d = all_data.get(tk)
        if not d:
            return 0
        acts = d["actuals"]
        if len(acts) < 4:
            return sum(a["revenue"] for a in acts)
        return sum(a["revenue"] for a in acts[-4:])

    sorted_tickers = sorted(
        [tk for tk in TICKERS if tk in all_data],
        key=fy_revenue, reverse=True)
    n_tk = len(sorted_tickers)

    # Build unified column layout
    cols, fy_groups, qmap, fymap = _summary_columns(all_data)
    last_ci = cols[-1]["ci"] if cols else 4

    # Build ticker-sheet column maps: for each ticker, map (year, quarter) -> ticker sheet column letter
    tk_col_maps = {}
    for tk in sorted_tickers:
        d = all_data[tk]
        tk_map = {}
        # Build the same column layout the ticker sheet uses
        tk_cols, tk_qmap, tk_fymap, _ = build_columns(d["actuals"], d["projections"], d["estimates"])
        for c in tk_cols:
            if c["t"] == "q":
                tk_map[(c["y"], c["q"])] = get_column_letter(c["ci"])
            elif c["t"] == "fy":
                tk_map[(c["y"], 0)] = get_column_letter(c["ci"])
        tk_col_maps[tk] = tk_map

    # ── Column widths ──
    ws.column_dimensions["A"].width = 8   # Ticker
    ws.column_dimensions["B"].width = 9   # Beat %
    ws.column_dimensions["C"].width = 12  # Momentum
    for c in cols:
        ws.column_dimensions[get_column_letter(c["ci"])].width = 11

    # ════════════════════════════════════════════════════════════════════════
    #  ROW 1: Title  |  ROW 2: Period headers
    # ════════════════════════════════════════════════════════════════════════

    ws.cell(row=1, column=1, value="KINETIC REVENUE MODEL").font = F_TITLE
    ws.cell(row=1, column=3,
            value=f"Generated {datetime.now().strftime('%B %d, %Y %H:%M')}").font = F_SUBTITLE

    # Fixed column headers
    ws.cell(row=2, column=1, value="Ticker").font = F_BOLD
    ws.cell(row=2, column=2, value="Beat %").font = F_BOLD
    ws.cell(row=2, column=3, value="Momentum").font = F_BOLD

    # Period headers
    for c in cols:
        cell = ws.cell(row=2, column=c["ci"], value=c["lbl"])
        cell.font = F_PERIOD
        cell.alignment = A_C

    # ════════════════════════════════════════════════════════════════════════
    #  FY GROUP BORDERS — applied after all sections are laid out
    # ════════════════════════════════════════════════════════════════════════

    # Helper to write a section of company rows + sum + optional avg
    def write_section(start_row, label, sorted_tks, cols, tk_col_maps,
                      tk_row, is_formula_row=False, is_pct=False,
                      ref_row=None, ref_row_ya=None, ref_row_pq=None,
                      green_red=False, consensus_blank_historical=False):
        """Write a section: header row, one row per ticker, sum row.

        tk_row: row number on the ticker sheet to reference
        ref_row/ref_row_ya/ref_row_pq: for formula rows (% YoY etc), the summary rows to reference
        """
        r = start_row
        # Section header
        ws.cell(row=r, column=1, value=label).font = F_BOLD
        r += 1

        first_tk_row = r
        for tk in sorted_tks:
            d = all_data[tk]
            tk_map = tk_col_maps[tk]

            # Fixed columns
            ws.cell(row=r, column=1, value=tk).font = F_DATA
            bc = d["beat_cadence"]
            if bc:
                c = ws.cell(row=r, column=2, value=bc["selected_beat_pct"] / 100)
                c.font = F_BLUE
                c.number_format = N_PCT
            ml = d["momentum_label"].upper()
            c = ws.cell(row=r, column=3, value=ml)
            c.font = F_MOM_GREEN if ml == "ACCELERATING" else (F_MOM_RED if ml == "DECELERATING" else F_DATA)

            # Data columns
            actual_yqs = set()
            for a in d["actuals"]:
                ad = datetime.strptime(a["period"], "%Y-%m-%d")
                actual_yqs.add((ad.year, quarter_from_date(a["period"])))

            for col in cols:
                ci = col["ci"]
                yq = (col["y"], col["q"]) if col["t"] == "q" else (col["y"], 0)
                tk_col_letter = tk_map.get(yq)

                if consensus_blank_historical and col["t"] == "q" and yq in actual_yqs:
                    continue  # blank for reported quarters in consensus section

                if not is_formula_row:
                    # Direct reference to ticker sheet cell
                    if tk_col_letter:
                        f = f"='{tk}'!{tk_col_letter}{tk_row}"
                        cell = ws.cell(row=r, column=ci, value=f)
                        cell.font = F_IT if is_pct else F_DATA
                        cell.number_format = N_PCT if is_pct else N_REV
                        cell.alignment = A_R
                else:
                    # Formula row (% YoY, % QoQ, $ YoY, $ QoQ) computed from summary rows
                    if ref_row_ya is not None and col["t"] == "q":
                        # % YoY or $ YoY — need same quarter prior year
                        ya_ci = qmap.get((col["y"] - 1, col["q"]))
                        if ya_ci:
                            if is_pct:
                                f = f'=IFERROR(({cr(ref_row, ci)}-{cr(ref_row, ya_ci)})/{cr(ref_row, ya_ci)},"-")'
                            else:
                                f = f"={cr(ref_row, ci)}-{cr(ref_row, ya_ci)}"
                            cell = ws.cell(row=r, column=ci, value=f)
                            cell.font = F_IT
                            cell.number_format = N_PCT if is_pct else N_DLR
                            cell.alignment = A_R
                    elif ref_row_ya is not None and col["t"] == "fy":
                        pfy = fymap.get(col["y"] - 1)
                        if pfy:
                            if is_pct:
                                f = f'=IFERROR(({cr(ref_row, ci)}-{cr(ref_row, pfy)})/{cr(ref_row, pfy)},"-")'
                            else:
                                f = f"={cr(ref_row, ci)}-{cr(ref_row, pfy)}"
                            cell = ws.cell(row=r, column=ci, value=f)
                            cell.font = F_IT
                            cell.number_format = N_PCT if is_pct else N_DLR
                            cell.alignment = A_R
                    elif ref_row_pq is not None and col["t"] == "q":
                        # % QoQ or $ QoQ
                        py, pq = prev_q(col["y"], col["q"])
                        pci = qmap.get((py, pq))
                        if pci:
                            if is_pct:
                                f = f'=IFERROR(({cr(ref_row, ci)}-{cr(ref_row, pci)})/{cr(ref_row, pci)},"-")'
                            else:
                                f = f"={cr(ref_row, ci)}-{cr(ref_row, pci)}"
                            cell = ws.cell(row=r, column=ci, value=f)
                            cell.font = F_IT
                            cell.number_format = N_PCT if is_pct else N_DLR
                            cell.alignment = A_R

                    if green_red:
                        # Can't easily pre-determine sign for formulas, skip color here
                        pass

            r += 1

        last_tk_row = r - 1

        # Total Revenue row with top border
        TOP_B = Border(top=FY_BORDER)
        ws.cell(row=r, column=1, value="Total Revenue").font = F_BOLD
        for ci2 in range(1, last_ci + 1):
            cell = ws.cell(row=r, column=ci2)
            cell.border = Border(top=FY_BORDER, left=cell.border.left,
                                 right=cell.border.right, bottom=cell.border.bottom)
        for col in cols:
            ci = col["ci"]
            if col["t"] == "fy":
                fq = col["fqc"]
                f = f"=SUM({cr(r, fq[0])}:{cr(r, fq[-1])})"
            else:
                f = f"=SUM({cr(first_tk_row, ci)}:{cr(last_tk_row, ci)})"
            cell = ws.cell(row=r, column=ci, value=f)
            cell.font = F_BOLD
            cell.number_format = N_REV if not is_pct else N_PCT
            cell.alignment = A_R
        sum_row = r
        r += 1

        return r, first_tk_row, last_tk_row, sum_row

    # ════════════════════════════════════════════════════════════════════════
    #  SECTION 1: KIN BASE CASE REVENUE
    # ════════════════════════════════════════════════════════════════════════

    r = 3  # start after header rows

    # Revenue
    r, ftk, ltk, rev_sum = write_section(
        r, "Total Revenue ($M)", sorted_tickers, cols, tk_col_maps, tk_row=9)
    rev_first_tk = ftk

    # Avg % YoY Growth row
    ws.cell(row=r, column=1, value="Avg % YoY Growth").font = F_IT
    for col in cols:
        ci = col["ci"]
        if col["t"] == "q":
            ya = qmap.get((col["y"] - 1, col["q"]))
            if ya:
                f = f'=IFERROR(({cr(rev_sum, ci)}-{cr(rev_sum, ya)})/{cr(rev_sum, ya)},"-")'
                cell = ws.cell(row=r, column=ci, value=f)
                cell.font = F_IT
                cell.number_format = N_PCT
                cell.alignment = A_R
        elif col["t"] == "fy":
            pfy = fymap.get(col["y"] - 1)
            if pfy:
                f = f'=IFERROR(({cr(rev_sum, ci)}-{cr(rev_sum, pfy)})/{cr(rev_sum, pfy)},"-")'
                cell = ws.cell(row=r, column=ci, value=f)
                cell.font = F_IT
                cell.number_format = N_PCT
                cell.alignment = A_R
    r += 2  # empty row

    # % YoY per company
    ws.cell(row=r, column=1, value="% YoY").font = F_BOLD
    r += 1
    yoy_first = r
    for tk in sorted_tickers:
        ws.cell(row=r, column=1, value=tk).font = F_DATA
        tk_map = tk_col_maps[tk]
        for col in cols:
            ci = col["ci"]
            yq = (col["y"], col["q"]) if col["t"] == "q" else (col["y"], 0)
            tk_rev_row = rev_first_tk + sorted_tickers.index(tk)
            if col["t"] == "q":
                ya = qmap.get((col["y"] - 1, col["q"]))
                if ya:
                    f = f'=IFERROR(({cr(tk_rev_row, ci)}-{cr(tk_rev_row, ya)})/{cr(tk_rev_row, ya)},"-")'
                    cell = ws.cell(row=r, column=ci, value=f)
                    cell.font = F_IT
                    cell.number_format = N_PCT
                    cell.alignment = A_R
            elif col["t"] == "fy":
                pfy = fymap.get(col["y"] - 1)
                if pfy:
                    f = f'=IFERROR(({cr(tk_rev_row, ci)}-{cr(tk_rev_row, pfy)})/{cr(tk_rev_row, pfy)},"-")'
                    cell = ws.cell(row=r, column=ci, value=f)
                    cell.font = F_IT
                    cell.number_format = N_PCT
                    cell.alignment = A_R
        r += 1
    # Total row for YoY
    ws.cell(row=r, column=1, value="Total Revenue").font = F_BOLD
    _total_row_border(ws, r, last_ci)
    for col in cols:
        ci = col["ci"]
        if col["t"] == "q":
            ya = qmap.get((col["y"] - 1, col["q"]))
            if ya:
                f = f'=IFERROR(({cr(rev_sum, ci)}-{cr(rev_sum, ya)})/{cr(rev_sum, ya)},"-")'
                cell = ws.cell(row=r, column=ci, value=f)
                cell.font = F_BOLD
                cell.number_format = N_PCT
                cell.alignment = A_R
        elif col["t"] == "fy":
            pfy = fymap.get(col["y"] - 1)
            if pfy:
                f = f'=IFERROR(({cr(rev_sum, ci)}-{cr(rev_sum, pfy)})/{cr(rev_sum, pfy)},"-")'
                cell = ws.cell(row=r, column=ci, value=f)
                cell.font = F_BOLD
                cell.number_format = N_PCT
                cell.alignment = A_R
    r += 2

    # % QoQ per company
    ws.cell(row=r, column=1, value="% QoQ").font = F_BOLD
    r += 1
    for tk in sorted_tickers:
        ws.cell(row=r, column=1, value=tk).font = F_DATA
        tk_rev_row = rev_first_tk + sorted_tickers.index(tk)
        for col in cols:
            if col["t"] != "q":
                continue
            ci = col["ci"]
            py, pq = prev_q(col["y"], col["q"])
            pci = qmap.get((py, pq))
            if pci:
                f = f'=IFERROR(({cr(tk_rev_row, ci)}-{cr(tk_rev_row, pci)})/{cr(tk_rev_row, pci)},"-")'
                cell = ws.cell(row=r, column=ci, value=f)
                cell.font = F_IT
                cell.number_format = N_PCT
                cell.alignment = A_R
        r += 1
    ws.cell(row=r, column=1, value="Total Revenue").font = F_BOLD
    _total_row_border(ws, r, last_ci)
    for col in cols:
        if col["t"] != "q":
            continue
        ci = col["ci"]
        py, pq = prev_q(col["y"], col["q"])
        pci = qmap.get((py, pq))
        if pci:
            f = f'=IFERROR(({cr(rev_sum, ci)}-{cr(rev_sum, pci)})/{cr(rev_sum, pci)},"-")'
            cell = ws.cell(row=r, column=ci, value=f)
            cell.font = F_BOLD
            cell.number_format = N_PCT
            cell.alignment = A_R
    r += 2

    # $ YoY per company
    ws.cell(row=r, column=1, value="$ YoY").font = F_BOLD
    r += 1
    for tk in sorted_tickers:
        ws.cell(row=r, column=1, value=tk).font = F_DATA
        tk_rev_row = rev_first_tk + sorted_tickers.index(tk)
        for col in cols:
            ci = col["ci"]
            if col["t"] == "q":
                ya = qmap.get((col["y"] - 1, col["q"]))
                if ya:
                    cell = ws.cell(row=r, column=ci,
                                   value=f"={cr(tk_rev_row, ci)}-{cr(tk_rev_row, ya)}")
                    cell.font = F_IT
                    cell.number_format = N_DLR
                    cell.alignment = A_R
            elif col["t"] == "fy":
                pfy = fymap.get(col["y"] - 1)
                if pfy:
                    cell = ws.cell(row=r, column=ci,
                                   value=f"={cr(tk_rev_row, ci)}-{cr(tk_rev_row, pfy)}")
                    cell.font = F_IT
                    cell.number_format = N_DLR
                    cell.alignment = A_R
        r += 1
    ws.cell(row=r, column=1, value="Total Revenue").font = F_BOLD
    _total_row_border(ws, r, last_ci)
    for col in cols:
        ci = col["ci"]
        if col["t"] == "q":
            ya = qmap.get((col["y"] - 1, col["q"]))
            if ya:
                cell = ws.cell(row=r, column=ci, value=f"={cr(rev_sum, ci)}-{cr(rev_sum, ya)}")
                cell.font = F_BOLD
                cell.number_format = N_DLR
                cell.alignment = A_R
        elif col["t"] == "fy":
            pfy = fymap.get(col["y"] - 1)
            if pfy:
                cell = ws.cell(row=r, column=ci, value=f"={cr(rev_sum, ci)}-{cr(rev_sum, pfy)}")
                cell.font = F_BOLD
                cell.number_format = N_DLR
                cell.alignment = A_R
    r += 2

    # $ QoQ per company
    ws.cell(row=r, column=1, value="$ QoQ").font = F_BOLD
    r += 1
    for tk in sorted_tickers:
        ws.cell(row=r, column=1, value=tk).font = F_DATA
        tk_rev_row = rev_first_tk + sorted_tickers.index(tk)
        for col in cols:
            if col["t"] != "q":
                continue
            ci = col["ci"]
            py, pq = prev_q(col["y"], col["q"])
            pci = qmap.get((py, pq))
            if pci:
                cell = ws.cell(row=r, column=ci,
                               value=f"={cr(tk_rev_row, ci)}-{cr(tk_rev_row, pci)}")
                cell.font = F_IT
                cell.number_format = N_DLR
                cell.alignment = A_R
        r += 1
    ws.cell(row=r, column=1, value="Total Revenue").font = F_BOLD
    _total_row_border(ws, r, last_ci)
    for col in cols:
        if col["t"] != "q":
            continue
        ci = col["ci"]
        py, pq = prev_q(col["y"], col["q"])
        pci = qmap.get((py, pq))
        if pci:
            cell = ws.cell(row=r, column=ci, value=f"={cr(rev_sum, ci)}-{cr(rev_sum, pci)}")
            cell.font = F_BOLD
            cell.number_format = N_DLR
            cell.alignment = A_R
    r += 2

    # ════════════════════════════════════════════════════════════════════════
    #  SECTION 2: CONSENSUS REVENUE
    # ════════════════════════════════════════════════════════════════════════

    r, con_ftk, con_ltk, con_sum = write_section(
        r, "Consensus Revenue ($M)", sorted_tickers, cols, tk_col_maps,
        tk_row=15, consensus_blank_historical=True)
    r += 1  # empty row

    # Consensus % YoY
    ws.cell(row=r, column=1, value="% YoY").font = F_BOLD
    r += 1
    for tk in sorted_tickers:
        ws.cell(row=r, column=1, value=tk).font = F_DATA
        con_tk_row = con_ftk + sorted_tickers.index(tk)
        for col in cols:
            ci = col["ci"]
            if col["t"] == "q":
                ya = qmap.get((col["y"] - 1, col["q"]))
                if ya:
                    f = f'=IFERROR(({cr(con_tk_row, ci)}-{cr(con_tk_row, ya)})/{cr(con_tk_row, ya)},"-")'
                    cell = ws.cell(row=r, column=ci, value=f)
                    cell.font = F_IT
                    cell.number_format = N_PCT
                    cell.alignment = A_R
        r += 1
    r += 1

    # Consensus % QoQ
    ws.cell(row=r, column=1, value="% QoQ").font = F_BOLD
    r += 1
    for tk in sorted_tickers:
        ws.cell(row=r, column=1, value=tk).font = F_DATA
        con_tk_row = con_ftk + sorted_tickers.index(tk)
        for col in cols:
            if col["t"] != "q":
                continue
            ci = col["ci"]
            py, pq = prev_q(col["y"], col["q"])
            pci = qmap.get((py, pq))
            if pci:
                f = f'=IFERROR(({cr(con_tk_row, ci)}-{cr(con_tk_row, pci)})/{cr(con_tk_row, pci)},"-")'
                cell = ws.cell(row=r, column=ci, value=f)
                cell.font = F_IT
                cell.number_format = N_PCT
                cell.alignment = A_R
        r += 1
    r += 1

    # Consensus $ YoY
    ws.cell(row=r, column=1, value="$ YoY").font = F_BOLD
    r += 1
    for tk in sorted_tickers:
        ws.cell(row=r, column=1, value=tk).font = F_DATA
        con_tk_row = con_ftk + sorted_tickers.index(tk)
        for col in cols:
            ci = col["ci"]
            if col["t"] == "q":
                ya = qmap.get((col["y"] - 1, col["q"]))
                if ya:
                    cell = ws.cell(row=r, column=ci,
                                   value=f"={cr(con_tk_row, ci)}-{cr(con_tk_row, ya)}")
                    cell.font = F_IT
                    cell.number_format = N_DLR
                    cell.alignment = A_R
        r += 1
    r += 1

    # Consensus $ QoQ
    ws.cell(row=r, column=1, value="$ QoQ").font = F_BOLD
    r += 1
    for tk in sorted_tickers:
        ws.cell(row=r, column=1, value=tk).font = F_DATA
        con_tk_row = con_ftk + sorted_tickers.index(tk)
        for col in cols:
            if col["t"] != "q":
                continue
            ci = col["ci"]
            py, pq = prev_q(col["y"], col["q"])
            pci = qmap.get((py, pq))
            if pci:
                cell = ws.cell(row=r, column=ci,
                               value=f"={cr(con_tk_row, ci)}-{cr(con_tk_row, pci)}")
                cell.font = F_IT
                cell.number_format = N_DLR
                cell.alignment = A_R
        r += 1
    r += 2

    # ════════════════════════════════════════════════════════════════════════
    #  SECTION 3: % VARIANCE VS CONSENSUS
    # ════════════════════════════════════════════════════════════════════════

    ws.cell(row=r, column=1, value="% Variance vs Consensus").font = F_BOLD
    r += 1
    var_first = r
    for tk in sorted_tickers:
        ws.cell(row=r, column=1, value=tk).font = F_DATA
        tk_rev_row = rev_first_tk + sorted_tickers.index(tk)
        con_tk_row = con_ftk + sorted_tickers.index(tk)
        for col in cols:
            ci = col["ci"]
            f = f'=IFERROR(({cr(tk_rev_row, ci)}-{cr(con_tk_row, ci)})/{cr(con_tk_row, ci)},"-")'
            cell = ws.cell(row=r, column=ci, value=f)
            cell.font = F_IT
            cell.number_format = N_PCT
            cell.alignment = A_R
        r += 1
    # Sum row for variance
    ws.cell(row=r, column=1, value="Total Revenue").font = F_BOLD
    _total_row_border(ws, r, last_ci)
    for col in cols:
        ci = col["ci"]
        f = f'=IFERROR(({cr(rev_sum, ci)}-{cr(con_sum, ci)})/{cr(con_sum, ci)},"-")'
        cell = ws.cell(row=r, column=ci, value=f)
        cell.font = F_BOLD
        cell.number_format = N_PCT
        cell.alignment = A_R
    r += 2

    # ════════════════════════════════════════════════════════════════════════
    #  FAR RIGHT: IMPLIED Q+2 GUIDE
    # ════════════════════════════════════════════════════════════════════════

    guide_col = last_ci + 2  # empty column separator
    ws.column_dimensions[get_column_letter(guide_col)].width = 14
    ws.column_dimensions[get_column_letter(guide_col + 1)].width = 14
    ws.column_dimensions[get_column_letter(guide_col + 2)].width = 10

    ws.cell(row=2, column=guide_col, value="Impl Guide").font = F_PERIOD
    ws.cell(row=2, column=guide_col + 1, value="Consensus").font = F_PERIOD
    ws.cell(row=2, column=guide_col + 2, value="Gap %").font = F_PERIOD

    for idx, tk in enumerate(sorted_tickers):
        d = all_data[tk]
        gi = d.get("guide_inference")
        r_tk = rev_first_tk + idx

        if gi:
            # Find Q+2 column letter on the ticker sheet
            tk_map = tk_col_maps[tk]
            gi_d = datetime.strptime(gi["period"], "%Y-%m-%d")
            gi_yq = (gi_d.year, quarter_from_date(gi["period"]))
            tk_col_letter = tk_map.get(gi_yq)

            if tk_col_letter:
                # Implied guide — cross-sheet reference to ticker sheet row 23
                c = ws.cell(row=r_tk, column=guide_col,
                            value=f"='{tk}'!{tk_col_letter}23")
                c.font = F_DATA
                c.number_format = N_REV
                c.alignment = A_R

                # Consensus — cross-sheet reference to ticker sheet row 15
                c = ws.cell(row=r_tk, column=guide_col + 1,
                            value=f"='{tk}'!{tk_col_letter}15")
                c.font = F_DATA
                c.number_format = N_REV
                c.alignment = A_R

                # Gap % — formula referencing the two cells above
                gc = get_column_letter(guide_col)
                gcc = get_column_letter(guide_col + 1)
                c = ws.cell(row=r_tk, column=guide_col + 2,
                            value=f'=IFERROR(({gc}{r_tk}-{gcc}{r_tk})/{gcc}{r_tk},"-")')
                c.font = F_IT
                c.number_format = N_PCT
                c.alignment = A_R

    # ════════════════════════════════════════════════════════════════════════
    #  FY GROUP BORDERS
    # ════════════════════════════════════════════════════════════════════════

    for yr, grp in fy_groups.items():
        for row in range(2, r + 1):
            cell = ws.cell(row=row, column=grp["first_ci"])
            cell.border = Border(left=FY_BORDER, right=cell.border.right,
                                 top=cell.border.top, bottom=cell.border.bottom)
            if grp["fy_ci"]:
                cell = ws.cell(row=row, column=grp["fy_ci"])
                cell.border = Border(left=FY_BORDER, right=FY_BORDER,
                                     top=cell.border.top, bottom=cell.border.bottom)

    # ── Conditional formatting: green/red for guide gap % column ──
    from openpyxl.formatting.rule import FormulaRule
    gap_col_letter = get_column_letter(guide_col + 2)
    gap_range = f"{gap_col_letter}{rev_first_tk}:{gap_col_letter}{rev_first_tk + n_tk - 1}"
    green_font = Font(name="Times New Roman", size=9, italic=True, color=GREEN_TXT)
    red_font = Font(name="Times New Roman", size=9, italic=True, color=RED_TXT)
    ws.conditional_formatting.add(gap_range, FormulaRule(
        formula=[f'{gap_col_letter}{rev_first_tk}>0'], font=green_font))
    ws.conditional_formatting.add(gap_range, FormulaRule(
        formula=[f'{gap_col_letter}{rev_first_tk}<0'], font=red_font))

    # ── Freeze panes: first 3 columns and first 2 rows ──
    ws.freeze_panes = "D3"


# ── main ──────────────────────────────────────────────────────────────────

def main():
    # Ensure consensus overrides are applied before generating Excel
    from ingest import apply_consensus_overrides
    apply_consensus_overrides()

    print("Building data...")
    all_data = build_all()
    wb = Workbook()
    build_summary_sheet(wb, all_data)
    for tk in TICKERS:
        if tk in all_data:
            build_ticker_sheet(wb, tk, all_data[tk])
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "kinetic_revenue_model.xlsx")
    wb.save(path)
    print(f"Exported to {path}")


if __name__ == "__main__":
    main()
